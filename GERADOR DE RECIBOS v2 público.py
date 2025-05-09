import os
import xml.etree.ElementTree as ET
from tkinter import *
from tkinter import ttk, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import locale
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side

class AplicativoRecibo:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Recibos v2")
        self.root.geometry("1000x700")
        
        # Configurar locale para português
        try:
            locale.setlocale(locale.LC_ALL, 'pt_BR.utf8')
        except:
            locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil')
        
        # Arquivo XML
        self.xml_file = "funcionarios.xml"
        self.funcionarios = []
        
        # Carregar dados
        self.carregar_funcionarios()
        
        # Variáveis
        self.funcionario_selecionado = StringVar()
        self.empresa = StringVar(value="EMPRESA")
        self.cnpj = StringVar(value="01.234.567/0001-89")
        self.valor = StringVar()
        self.valor_extenso = StringVar()
        self.periodo = StringVar(value=datetime.now().strftime("%B/%Y").capitalize())
        self.cidade = StringVar(value="CIDADE")
        self.pagamento_checkbox = BooleanVar(value=False)
        
        # Variáveis para cadastro
        self.novo_nome = StringVar()
        self.novo_cnpj = StringVar()
        self.novo_salario = StringVar()
        self.novo_adiantamento = StringVar()
        self.novo_parcela = StringVar()
        
        # Variáveis para descontos
        self.descontos = []  # Lista para armazenar descontos
        self.frame_descontos = None  # Frame para os descontos
        
        # Criar abas
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=BOTH, expand=True)
        
        # Abas principais
        self.criar_aba_recibo()
        self.criar_aba_funcionarios()
        self.criar_aba_preview()
        
        # Atualizar combobox
        self.atualizar_combobox_funcionarios()

    def formatar_brasileiro(self, valor):
        """Formata um valor float ou string no padrão brasileiro (1.000,00)"""
        try:
            # Se for string, converte para float tratando o formato brasileiro
            if isinstance(valor, str):
                # Primeiro verifica se tem vírgula (separador decimal)
                if "," in valor:
                    # Remove pontos de milhar e substitui vírgula decimal por ponto
                    partes = valor.split(",")
                    parte_inteira = partes[0].replace(".", "")
                    valor = float(f"{parte_inteira}.{partes[1]}")
                else:
                    valor = float(valor.replace(".", ""))
            
            # Formata com separador de milhar e decimal com vírgula
            valor_str = f"{valor:,.2f}"
            inteiro, decimal = valor_str.split(".")
            inteiro = inteiro.replace(",", ".")
            return f"{inteiro},{decimal}"
        except:
            return "0,00"

    def numero_para_extenso(self, numero):
        """Converte um número para sua representação por extenso em português"""
        unidades = ["", "um", "dois", "três", "quatro", "cinco", 
                   "seis", "sete", "oito", "nove"]
        dez_a_dezenove = ["dez", "onze", "doze", "treze", "quatorze", 
                         "quinze", "dezesseis", "dezessete", "dezoito", "dezenove"]
        dezenas = ["vinte", "trinta", "quarenta", "cinquenta", 
                  "sessenta", "setenta", "oitenta", "noventa"]
        centenas = ["cento", "duzentos", "trezentos", "quatrocentos", 
                   "quinhentos", "seiscentos", "setecentos", "oitocentos", "novecentos"]
        
        try:
            # Tratar entrada como string (formato brasileiro 1.000,00)
            if isinstance(numero, str):
                # Remover pontos de milhar e substituir vírgula decimal por ponto
                numero = numero.replace(".", "").replace(",", ".")
            numero = float(numero)
        except:
            return "Valor inválido"
        
        if numero == 0:
            return "Zero reais"
        
        inteiro = int(numero)
        decimal = int(round((numero - inteiro) * 100))
        
        extenso = []
        
        # Tratar milhares
        if inteiro >= 1000:
            milhares = inteiro // 1000
            inteiro %= 1000
            
            if milhares == 1:
                extenso.append("mil")
            else:
                # Usar recursão para números complexos (ex: 2.300 → "dois mil e trezentos")
                parte_milhar = self.numero_para_extenso(str(milhares)).replace(" reais", "").replace(" real", "")
                extenso.append(parte_milhar)
                extenso.append("mil")
                if inteiro > 0 and inteiro < 100:
                    extenso.append("e")
        
        # Tratar centenas
        if inteiro >= 100:
            if inteiro == 100:
                extenso.append("cem")
            else:
                extenso.append(centenas[(inteiro // 100) - 1])
            inteiro %= 100
            if inteiro > 0:
                extenso.append("e")
        
        # Tratar dezenas e unidades
        if inteiro >= 20:
            extenso.append(dezenas[(inteiro // 10) - 2])
            inteiro %= 10
            if inteiro > 0:
                extenso.append("e")
                extenso.append(unidades[inteiro])
        elif 10 <= inteiro <= 19:
            extenso.append(dez_a_dezenove[inteiro - 10])
        elif 1 <= inteiro <= 9:
            extenso.append(unidades[inteiro])
        
        # Tratar centavos
        if decimal > 0:
            if len(extenso) > 0:
                extenso.append("e")
            if 10 <= decimal <= 19:
                extenso.append(dez_a_dezenove[decimal - 10])
            elif decimal >= 20:
                extenso.append(dezenas[(decimal // 10) - 2])
                if decimal % 10 > 0:
                    extenso.append("e")
                    extenso.append(unidades[decimal % 10])
            else:
                extenso.append(unidades[decimal])
            extenso.append("centavos")
        else:
            if len(extenso) > 0:
                extenso.append("reais")
        
        # Juntar tudo e capitalizar a primeira letra
        if extenso:
            extenso[0] = extenso[0].capitalize()
            texto = " ".join(extenso)
            
            # Correções gramaticais
            texto = texto.replace("um mil", "mil")  # Correto: "mil reais" em vez de "um mil reais"
            texto = texto.replace("  ", " ")       # Remove espaços duplos
            texto = texto.replace("e cento", "cento")
            texto = texto.replace("e cem", "cem")
            return texto
        
        return "Zero reais"

    def adicionar_desconto(self):
        # Criar um novo frame para o desconto
        frame_desconto = ttk.Frame(self.frame_descontos)
        frame_desconto.pack(fill=X, pady=5)
        
        # Variáveis para este desconto
        descricao = StringVar(value=f"Desconto {len(self.descontos)+1}")
        valor = StringVar(value="0,00")
        
        # Widgets para o desconto
        ttk.Label(frame_desconto, text="Descrição:").pack(side=LEFT, padx=5)
        ttk.Entry(frame_desconto, textvariable=descricao, width=20).pack(side=LEFT, padx=5)
        
        ttk.Label(frame_desconto, text="Valor (R$):").pack(side=LEFT, padx=5)
        ttk.Entry(frame_desconto, textvariable=valor, width=10).pack(side=LEFT, padx=5)
        
        # Botão para remover este desconto
        ttk.Button(frame_desconto, text="×", width=3, 
                  command=lambda: self.remover_desconto(frame_desconto)).pack(side=LEFT, padx=5)
        
        # Adicionar à lista de descontos
        self.descontos.append({
            'frame': frame_desconto,
            'descricao': descricao,
            'valor': valor
        })
        
        # Atualizar preview
        self.atualizar_preview()
    
    def remover_desconto(self, frame_desconto):
        # Encontrar e remover o desconto
        for desconto in self.descontos[:]:
            if desconto['frame'] == frame_desconto:
                self.descontos.remove(desconto)
                frame_desconto.destroy()
                break
        
        # Renumerar os descontos restantes
        for i, desconto in enumerate(self.descontos):
            desconto['descricao'].set(f"Desconto {i+1}")
        
        # Atualizar preview
        self.atualizar_preview()
    
    def carregar_funcionarios(self):
        if not os.path.exists(self.xml_file):
            # Criar XML inicial se não existir
            root = ET.Element("funcionarios")
            tree = ET.ElementTree(root)
            tree.write(self.xml_file)
        
        try:
            tree = ET.parse(self.xml_file)
            root = tree.getroot()
            
            self.funcionarios = []
            for func in root.findall('funcionario'):
                self.funcionarios.append({
                    'nome': func.find('nome').text,
                    'cnpj': func.find('cnpj').text if func.find('cnpj') is not None else '',
                    'salario': func.find('salario').text if func.find('salario') is not None else '0',
                    'adiantamento': func.find('adiantamento').text if func.find('adiantamento') is not None else '0',
                    'parcela_extra': func.find('parcela_extra').text if func.find('parcela_extra') is not None else '0'
                })
        except:
            self.funcionarios = []
    
    def salvar_funcionarios(self):
        root = ET.Element("funcionarios")
        
        for func in self.funcionarios:
            funcionario = ET.SubElement(root, "funcionario")
            ET.SubElement(funcionario, "nome").text = func['nome']
            ET.SubElement(funcionario, "cnpj").text = func['cnpj']
            ET.SubElement(funcionario, "salario").text = func['salario']
            ET.SubElement(funcionario, "adiantamento").text = func['adiantamento']
            ET.SubElement(funcionario, "parcela_extra").text = func['parcela_extra']
        
        tree = ET.ElementTree(root)
        tree.write(self.xml_file)
    
    def criar_aba_recibo(self):
        aba = ttk.Frame(self.notebook)
        self.notebook.add(aba, text="Emitir Recibo")
        
        mainframe = ttk.Frame(aba, padding="20")
        mainframe.pack(fill=BOTH, expand=True)
        
        # Título
        ttk.Label(mainframe, text="Emitir Recibo de Quitação", font=('Arial', 14, 'bold')).grid(column=0, row=0, columnspan=2, pady=10)

        # Variável para o modo de pagamento
        self.modo_pagamento = StringVar(value="Adiantamento")
        
        # Frame para seleção do modo
        frame_modo = ttk.Frame(mainframe)
        frame_modo.grid(column=0, row=1, columnspan=2, sticky=EW, pady=5)
        
        ttk.Label(frame_modo, text="Tipo:").pack(side=LEFT, padx=5)
        
        # Radiobuttons para seleção do modo
        ttk.Radiobutton(
            frame_modo, 
            text="Adiantamento",
            variable=self.modo_pagamento,
            value="Adiantamento",
            command=self.atualizar_valores_pagamento
        ).pack(side=LEFT, padx=5)
        
        ttk.Radiobutton(
            frame_modo, 
            text="Pagamento",
            variable=self.modo_pagamento,
            value="Pagamento",
            command=self.atualizar_valores_pagamento
        ).pack(side=LEFT, padx=5)
        
        # Checkbox para pagamento completo
        frame_pagamento = ttk.Frame(mainframe)
        frame_pagamento.grid(column=1, row=2, columnspan=2, sticky=W, pady=5)
        
        ttk.Checkbutton(
            frame_pagamento,
            text="Pagamento completo",
            variable=self.pagamento_checkbox,
            command=self.atualizar_valores_pagamento,
            
        ).pack(side = "left")

        # Seleção de funcionário
        ttk.Label(mainframe, text="Funcionário:").grid(column=0, row=2, sticky=W, pady=5)
        self.cb_funcionarios = ttk.Combobox(mainframe, textvariable=self.funcionario_selecionado, state="readonly")
        self.cb_funcionarios.grid(column=0, row=3, columnspan=2, sticky=EW, pady=5)
        self.cb_funcionarios.bind("<<ComboboxSelected>>", self.selecionar_funcionario)
        
        # Dados da empresa
        ttk.Label(mainframe, text="Empresa:").grid(column=0, row=4, sticky=W, pady=5)
        ttk.Entry(mainframe, width=40, textvariable=self.empresa).grid(column=0, row=5, columnspan=2, sticky=EW)
        
        ttk.Label(mainframe, text="CNPJ:").grid(column=0, row=6, sticky=W, pady=5)
        ttk.Entry(mainframe, width=25, textvariable=self.cnpj).grid(column=0, row=7, sticky=W)
        
        # Valor e período
        ttk.Label(mainframe, text="Valor (R$):").grid(column=0, row=8, sticky=W, pady=5)
        ttk.Entry(mainframe, width=15, textvariable=self.valor).grid(column=0, row=9, sticky=W)
        self.valor.trace_add("write", self.atualizar_valor_extenso)
        
        ttk.Label(mainframe, text="Valor por extenso:").grid(column=0, row=10, sticky=W, pady=5)
        ttk.Entry(mainframe, width=40, textvariable=self.valor_extenso).grid(column=0, row=11, columnspan=2, sticky=EW)
        
        ttk.Label(mainframe, text="Período:").grid(column=0, row=12, sticky=W, pady=5)
        ttk.Entry(mainframe, width=20, textvariable=self.periodo).grid(column=0, row=13, sticky=W)
        
        ttk.Label(mainframe, text="Cidade:").grid(column=0, row=14, sticky=W, pady=5)
        ttk.Entry(mainframe, width=20, textvariable=self.cidade).grid(column=0, row=15, sticky=W)
        
        # Frame para descontos
        ttk.Label(mainframe, text="Descontos:").grid(column=0, row=16, sticky=W, pady=5)
        
        self.frame_descontos = ttk.Frame(mainframe)
        self.frame_descontos.grid(column=0, row=17, columnspan=2, sticky=EW)
        
        # Botão para adicionar desconto
        ttk.Button(mainframe, text="+ Adicionar Desconto", 
                  command=self.adicionar_desconto).grid(column=0, row=18, pady=10, sticky=W)
        
        # Botão de gerar recibo
        ttk.Button(mainframe, text="Gerar Recibo", command=self.gerar_recibo).grid(column=1, row=18, pady=10, sticky=E)

    def atualizar_valores_pagamento(self):
        """Atualiza os valores automaticamente quando o modo de pagamento é alterado"""
        nome = self.funcionario_selecionado.get()
        if not nome:
            return
        
        # Encontrar o funcionário selecionado
        funcionario = None
        for func in self.funcionarios:
            if func['nome'] == nome:
                funcionario = func
                break
        
        if not funcionario:
            return
        
        try:
            # Converter valores para float
            salario = float(funcionario['salario'].replace(".", "").replace(",", "."))
            adiantamento = float(funcionario['adiantamento'].replace(".", "").replace(",", "."))
            
            if self.modo_pagamento.get() == "Pagamento" and self.pagamento_checkbox.get():
                # Modo pagamento com checkbox marcado: calcular salário - adiantamento
                valor_liquido = salario
                self.valor.set(self.formatar_brasileiro(valor_liquido))
            else:
                # Modo adiantamento ou pagamento sem checkbox: usar valor do adiantamento
                self.valor.set(funcionario['adiantamento'])
            
            # Atualizar valor por extenso
            valor_num = float(self.valor.get().replace(".", "").replace(",", "."))
            self.valor_extenso.set(self.numero_para_extenso(valor_num))
            
        except ValueError as e:
            print(f"Erro ao converter valores: {e}")
            messagebox.showerror("Erro", "Valores numéricos inválidos no cadastro do funcionário!")
        
        self.atualizar_preview()
    
    def atualizar_valor_extenso(self, *args):
        try:
            total_float = self.calcular_total(retornar_float=True)
            self.valor_extenso.set(self.numero_para_extenso(total_float))
        except ValueError:
            self.valor_extenso.set("Valor inválido")
        self.atualizar_preview()
    
    def criar_aba_funcionarios(self):
        aba = ttk.Frame(self.notebook)
        self.notebook.add(aba, text="Gerenciar Funcionários")
        
        mainframe = ttk.Frame(aba, padding="20")
        mainframe.pack(fill=BOTH, expand=True)
        
        # Título
        ttk.Label(mainframe, text="Cadastro de Funcionários", font=('Arial', 14, 'bold')).grid(column=0, row=0, columnspan=3, pady=10)
        
        # Formulário de cadastro
        ttk.Label(mainframe, text="Nome:").grid(column=0, row=1, sticky=W, pady=5)
        ttk.Entry(mainframe, textvariable=self.novo_nome).grid(column=0, row=2, sticky=EW, padx=5)
        
        ttk.Label(mainframe, text="CPF/CNPJ:").grid(column=1, row=1, sticky=W, pady=5)
        ttk.Entry(mainframe, textvariable=self.novo_cnpj).grid(column=1, row=2, sticky=EW, padx=5)
        
        ttk.Label(mainframe, text="Salário (R$):").grid(column=2, row=1, sticky=W, pady=5)
        ttk.Entry(mainframe, textvariable=self.novo_salario).grid(column=2, row=2, sticky=EW, padx=5)
        
        ttk.Label(mainframe, text="Adiantamento Padrão (R$):").grid(column=3, row=1, sticky=W, pady=5)
        ttk.Entry(mainframe, textvariable=self.novo_adiantamento).grid(column=3, row=2, sticky=EW, padx=5)

        ttk.Label(mainframe, text="Parcela Extra Padrão (R$):").grid(column=4, row=1, sticky=W, pady=5)
        ttk.Entry(mainframe, textvariable=self.novo_parcela).grid(column=4, row=2, sticky=EW, padx=5)
        
        # Botões
        btn_frame = ttk.Frame(mainframe)
        btn_frame.grid(column=0, row=3, columnspan=4, pady=10)
        
        ttk.Button(btn_frame, text="Adicionar", command=self.adicionar_funcionario).pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="Editar", command=self.editar_funcionario).pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="Excluir", command=self.excluir_funcionario).pack(side=LEFT, padx=5)
        
        # Lista de funcionários
        columns = ("Nome", "CPF/CNPJ", "Salário", "Adiantamento", "Parcela Extra")
        self.tree_funcionarios = ttk.Treeview(mainframe, columns=columns, show="headings", height=15)
        
        for col in columns:
            self.tree_funcionarios.heading(col, text=col)
            self.tree_funcionarios.column(col, width=150, anchor=W)
        
        self.tree_funcionarios.grid(column=0, row=5, columnspan=5, sticky=NSEW, pady=10)
        self.tree_funcionarios.bind("<<TreeviewSelect>>", self.selecionar_funcionario_lista)
        
        # Atualizar lista
        self.atualizar_lista_funcionarios()
    
    def criar_aba_preview(self):
        aba = ttk.Frame(self.notebook)
        self.notebook.add(aba, text="Pré-visualização")
        
        mainframe = ttk.Frame(aba)
        mainframe.pack(fill=BOTH, expand=True)
        
        # Canvas e barra de rolagem
        canvas = Canvas(mainframe)
        scrollbar = ttk.Scrollbar(mainframe, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Pré-visualização
        self.preview = Text(scrollable_frame, width=100, height=35, wrap=WORD, font=('Arial', 11))
        self.preview.pack(padx=20, pady=20, fill=BOTH, expand=True)
        self.preview.config(state=DISABLED)
        
        # Atualizar preview inicial
        self.atualizar_preview()
    
    def atualizar_combobox_funcionarios(self):
        nomes = [f['nome'] for f in self.funcionarios]
        self.cb_funcionarios['values'] = nomes
        if nomes:
            self.cb_funcionarios.current(0)
            self.selecionar_funcionario()
    
    def atualizar_lista_funcionarios(self):
        self.tree_funcionarios.delete(*self.tree_funcionarios.get_children())
        for func in self.funcionarios:
            try:
                # Converter valores com vírgula para float
                salario = float(func['salario'].replace(".", "").replace(",", "."))
                adiantamento = float(func['adiantamento'].replace(".", "").replace(",", "."))
                parcela_extra = float(func['parcela_extra'].replace(".", "").replace(",", "."))
                
                self.tree_funcionarios.insert("", "end", values=(
                    func['nome'],
                    func['cnpj'],
                    f"R$ {salario:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                    f"R$ {adiantamento:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                    f"R$ {parcela_extra:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                ))
            except ValueError as e:
                print(f"Erro ao converter valores numéricos: {e}")
                # Insere os valores originais se houver erro de conversão
                self.tree_funcionarios.insert("", "end", values=(
                    func['nome'],
                    func['cnpj'],
                    f"R$ {func['salario']}",
                    f"R$ {func['adiantamento']}",
                    f"R$ {func['parcela_extra']}"
                ))
    
    def selecionar_funcionario(self, event=None):
        nome = self.funcionario_selecionado.get()
        if nome:
            self.atualizar_valores_pagamento()

    def selecionar_funcionario_lista(self, event):
        """Preenche os campos de edição quando seleciona um funcionário na lista"""
        selection = self.tree_funcionarios.selection()
        if selection:
            item = self.tree_funcionarios.item(selection[0])
            valores = item['values']
                
            self.novo_nome.set(valores[0])
            self.novo_cnpj.set(valores[1])
                
            # Formata os valores monetários para edição
            try:
                salario = valores[2].replace("R$ ", "").replace(".", "").replace(",", ".")
                self.novo_salario.set(salario)
                    
                adiantamento = valores[3].replace("R$ ", "").replace(".", "").replace(",", ".")
                self.novo_adiantamento.set(adiantamento)

                parcela_extra = valores[4].replace("R$ ", "").replace(".", "").replace(",", ".")
                self.novo_parcela.set(parcela_extra)
            except:
                self.novo_salario.set(valores[2].replace("R$ ", ""))
                self.novo_adiantamento.set(valores[3].replace("R$ ", ""))
                self.novo_parcela.set(valores[4].replace("R$ ", ""))
    
    def adicionar_funcionario(self):
        if not self.novo_nome.get():
            messagebox.showerror("Erro", "Informe o nome do funcionário!")
            return
        
        try:
            # Converter valores para formato com ponto antes de salvar
            salario = self.novo_salario.get().replace(".", "").replace(",", ".")
            adiantamento = self.novo_adiantamento.get().replace(".", "").replace(",", ".")
            parcela_extra = self.novo_parcela.get().replace(".", "").replace(",", ".")
            
            # Validar se são números válidos
            float(salario)
            float(adiantamento)
            float(parcela_extra)
            
            novo_func = {
                'nome': self.novo_nome.get(),
                'cnpj': self.novo_cnpj.get(),
                'salario': salario,
                'adiantamento': adiantamento,
                'parcela_extra': parcela_extra
            }
            
            self.funcionarios.append(novo_func)
            self.salvar_funcionarios()
            self.atualizar_lista_funcionarios()
            self.atualizar_combobox_funcionarios()
            
            self.limpar_campos_funcionario()
            messagebox.showinfo("Sucesso", "Funcionário adicionado com sucesso!")
        
        except ValueError:
            messagebox.showerror("Erro", "Valores numéricos inválidos! Use o formato 1234,56 ou 1234.56")
    
    def editar_funcionario(self):
        item = self.tree_funcionarios.selection()
        if not item:
            messagebox.showerror("Erro", "Selecione um funcionário para editar!")
            return
        
        nome_original = self.tree_funcionarios.item(item, 'values')[0]
        
        for i, func in enumerate(self.funcionarios):
            if func['nome'] == nome_original:
                self.funcionarios[i] = {
                    'nome': self.novo_nome.get(),
                    'cnpj': self.novo_cnpj.get(),
                    'salario': self.novo_salario.get() or '0',
                    'adiantamento': self.novo_adiantamento.get() or '0',
                    'parcela_extra':self.novo_parcela.get() or '0'
                }
                break
        
        self.salvar_funcionarios()
        self.atualizar_lista_funcionarios()
        self.atualizar_combobox_funcionarios()
        
        self.limpar_campos_funcionario()
        messagebox.showinfo("Sucesso", "Funcionário atualizado com sucesso!")
    
    def excluir_funcionario(self):
        item = self.tree_funcionarios.selection()
        if not item:
            messagebox.showerror("Erro", "Selecione um funcionário para excluir!")
            return
        
        nome = self.tree_funcionarios.item(item, 'values')[0]
        
        if messagebox.askyesno("Confirmar", f"Tem certeza que deseja excluir {nome}?"):
            self.funcionarios = [f for f in self.funcionarios if f['nome'] != nome]
            self.salvar_funcionarios()
            self.atualizar_lista_funcionarios()
            self.atualizar_combobox_funcionarios()
            self.limpar_campos_funcionario()
            messagebox.showinfo("Sucesso", "Funcionário excluído com sucesso!")
    
    def limpar_campos_funcionario(self):
        self.novo_nome.set("")
        self.novo_cnpj.set("")
        self.novo_salario.set("")
        self.novo_adiantamento.set("")
        self.novo_parcela.set("")
    
    def calcular_total(self, retornar_float=False):
        try:
            # Converter valor principal
            valor_principal = float(self.formatar_brasileiro(self.valor.get() or "0").replace(".", "").replace(",", "."))
            
            # Obter valores adicionais do funcionário selecionado
            adiantamento = 0.0
            parcela_extra = 0.0
            for func in self.funcionarios:
                if func['nome'] == self.funcionario_selecionado.get():
                    try:
                        adiantamento = float(func['adiantamento'].replace(".", "").replace(",", "."))
                        parcela_extra = float(func['parcela_extra'].replace(".", "").replace(",", "."))
                    except:
                        adiantamento = 0.0
                        parcela_extra = 0.0
                    break
            
            # Lógica de cálculo
            if self.pagamento_checkbox.get():
                # Modo quitação total: subtrai adiantamento e soma parcela extra
                valor_principal = valor_principal - adiantamento + parcela_extra
            
            # Subtrair descontos manuais
            for desconto in self.descontos:
                try:
                    valor_desconto = float(desconto['valor'].get().replace(".", "").replace(",", "."))
                    valor_principal -= valor_desconto
                except ValueError:
                    continue
            
            # Retorna conforme o parâmetro
            if retornar_float:
                return valor_principal
            else:
                return self.formatar_brasileiro(valor_principal)
                
        except Exception as e:
            print(f"Erro no cálculo do total: {e}")
            return 0.0 if retornar_float else "0,00"
    
    def atualizar_preview(self):
        nome_funcionario = self.funcionario_selecionado.get() or "NOME DO FUNCIONÁRIO"
        cpf_funcionario = ""
        
        for func in self.funcionarios:
            if func['nome'] == nome_funcionario:
                cpf_funcionario = func['cnpj']
                break
        
        # Determina o texto do título baseado no modo selecionado
        tipo_recibo = "HONORÁRIOS" if self.modo_pagamento.get() == "Pagamento" else "ADIANTAMENTO"
        
        # Adicionar descontos ao preview
        texto_descontos = ""
        
        for desconto in self.descontos:
            texto_descontos += f"\n{desconto['descricao'].get()} R$ {self.formatar_brasileiro(desconto['valor'].get())}"
        
        total_float = self.calcular_total(retornar_float=True)
        valor_extenso_total = self.numero_para_extenso(total_float)

        texto = f"""RECIBO DE QUITAÇÃO DE {tipo_recibo}

    Declaro que recebi da Empresa {self.empresa.get()} 
    CNPJ: {self.cnpj.get()}, a importância de R$ {self.formatar_brasileiro(self.valor.get() or '0')} 
    ({valor_extenso_total or 'Zero reais'}) referente a {self.modo_pagamento.get().lower()} para serviços prestados, 
    no período de {self.periodo.get()}, pelo que dou total quitação nada tenho a reclamar no futuro.

    VALORES DISCRIMINADOS A RECEBER

    {tipo_recibo.capitalize()} R$ {self.formatar_brasileiro(self.valor.get() or '0')}{texto_descontos}

    Total a pagar: R$ {self.formatar_brasileiro(self.calcular_total())}

    {self.cidade.get()}, __/__/____

    ________________
    {nome_funcionario}
    CPF/CNPJ {cpf_funcionario}"""
        
        self.preview.config(state=NORMAL)
        self.preview.delete(1.0, END)
        self.preview.insert(1.0, texto)
        self.preview.config(state=DISABLED)
    
    def gerar_recibo(self):
        if not self.funcionario_selecionado.get():
            messagebox.showerror("Erro", "Selecione um funcionário!")
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Recibo"
            row_desc = 22  # Linha inicial para descontos

            # Determina o texto do título baseado no modo selecionado
            tipo_recibo = "PAGAMENTO" if self.modo_pagamento.get() == "Pagamento" else "ADIANTAMENTO"
            
            # Estilos
            # Estilo de borda espessa
            borda_espessa = Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )
            formato_moeda = 'R$ #,##0.00'
            fill_amarelo = PatternFill(start_color="FFFF00", fill_type="solid")
            fill_vermelho = PatternFill(start_color="F4B084", fill_type="solid")
            titulo_1 = Font(name= 'Calibri', size=20, color='00B050', bold=True)
            fonte_roman = Font(name= 'Times New Roman', size= 12)
            roman_negrito = Font(name= 'Times New Roman', size= 14, bold=True)
            roman_negrito_sublinhado = Font(name= 'Times New Roman', size= 14, bold=True, underline='single')
            fonte_titulo = Font(size=14, bold=True)
            fonte_normal = Font(size=12)
            fonte_14 = Font(name= 'Calibri',size=14)
            fonte_negrito = Font(size=14, bold=True)
            negrito_11 = Font(size=11, bold=True)
            negrito_14 = Font(size=14, bold=True)
            alinhamento_centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
            alinhamento_esquerda = Alignment(horizontal='left', vertical='center', wrap_text=True)
            alinhamento_direita = Alignment(horizontal='right', vertical='center', wrap_text=True)

            # Obter o valor da parcela extra do funcionário selecionado
            parcela_extra = "0,00"
            for func in self.funcionarios:
                if func['nome'] == self.funcionario_selecionado.get():
                    parcela_extra = func['parcela_extra']
                    break

            # Título
            ws.merge_cells('A1:H2')
            ws['A1'] = f"RECIBO DE QUITAÇÃO DE {tipo_recibo}"
            ws['A1'].font = titulo_1
            ws['A1'].alignment = alinhamento_centro
            for row in ws['A1:H2']:
                for cell in row:
                    cell.border = borda_espessa

            # Corpo

            total_float = self.calcular_total(retornar_float=True)
            valor_extenso_total = self.numero_para_extenso(total_float)
            
            texto_corpo = (
                f"Declaro que recebi da Empresa {self.empresa.get()} "
                f"CNPJ: {self.cnpj.get()}, a importância de R$ {self.calcular_total()} "
                f"({valor_extenso_total}) referente ao {self.modo_pagamento.get().lower()} para serviços prestados, "
                f"no período de {self.periodo.get()}, pelo que dou total quitação nada tenho a reclamar no futuro."
            )

            ws.merge_cells('A6:H9')
            ws['A6'] = texto_corpo
            ws['A6'].font = fonte_roman
            ws['A6'].alignment = alinhamento_centro

            # Valores
            ws.merge_cells('A12:H12')
            ws['A12'] = "VALORES DISCRIMINADOS A RECEBER"
            ws['A12'].font = fonte_negrito
            ws['A12'].alignment = alinhamento_centro
            ws['A12'].fill = fill_amarelo

            ws['A16'] = f"{tipo_recibo.capitalize()}"
            ws['A16'].font = roman_negrito
            ws['E16'] = f"{self.formatar_brasileiro(self.valor.get())}"
            ws['E16'].font = fonte_negrito
            ws['E16'].number_format = formato_moeda
            ws['E16'].alignment = alinhamento_direita
            ws['E16'].number_format = 'R$ #,##0.00'

            # Seção da parcela extra
            if self.modo_pagamento.get() == "Pagamento" and self.pagamento_checkbox.get():
                ws['A17'] = "PARCELA EXTRA"
                ws['A17'].font = roman_negrito
                ws['E17'] = float(parcela_extra.replace(".", "").replace(",", "."))
                ws['E17'].number_format = 'R$ #,##0.00'
                ws['E17'].font = fonte_negrito
                ws['E17'].alignment = alinhamento_direita
            
            if self.pagamento_checkbox.get():  # Se estiver no modo de quitação total
                # Obter o valor do adiantamento do funcionário selecionado
                adiantamento = 0.0
                for func in self.funcionarios:
                    if func['nome'] == self.funcionario_selecionado.get():
                        try:
                            adiantamento = float(func['adiantamento'].replace(".", "").replace(",", "."))
                        except:
                            adiantamento = 0.0
                        break
                
                # Cabeçalho de descontos
                ws['A19'] = "DESCONTOS"
                ws['A19'].font = roman_negrito_sublinhado
                ws['A19'].fill = fill_vermelho
                ws['B19'].fill = fill_vermelho
                
                # Adiciona o adiantamento como primeiro desconto
                ws['A22'] = "ADIANTAMENTO"
                ws['E22'] = adiantamento * -1  # Valor negativo
                ws['E22'].number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
                ws['E22'].font = roman_negrito
                ws['A22'].font = roman_negrito
                row_desc = 23
                
                # Adiciona outros descontos se existirem
                if self.descontos:
                    for desconto in self.descontos:
                        try:
                            valor = float(desconto['valor'].get().replace(".", "").replace(",", "."))
                            if valor != 0:
                                ws[f'A{row_desc}'] = desconto['descricao'].get()
                                ws[f'E{row_desc}'] = valor * -1
                                ws[f'E{row_desc}'].number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
                                ws[f'E{row_desc}'].font = roman_negrito
                                ws[f'A{row_desc}'].font = roman_negrito
                                row_desc += 1
                        except ValueError:
                            continue
            else:
                # Modo normal sem quitação total - apenas descontos manuais
                if self.descontos:
                    ws['A19'] = "DESCONTOS"
                    ws['A19'].font = roman_negrito_sublinhado
                    ws['A19'].fill = fill_vermelho
                    ws['B19'].fill = fill_vermelho
                    
                    row_desc = 20
                    for desconto in self.descontos:
                        try:
                            valor = float(desconto['valor'].get().replace(".", "").replace(",", "."))
                            if valor != 0:
                                ws[f'A{row_desc}'] = desconto['descricao'].get()
                                ws[f'E{row_desc}'] = valor * -1
                                ws[f'E{row_desc}'].number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
                                ws[f'E{row_desc}'].font = roman_negrito
                                ws[f'A{row_desc}'].font = roman_negrito
                                row_desc += 1
                        except ValueError:
                            continue              

            # Total
            ws[f'A{row_desc}'] = "TOTAL A PAGAR:"
            ws[f'A{row_desc}'].font = roman_negrito_sublinhado
            total_float = self.calcular_total(retornar_float=True)
            ws[f'E{row_desc}'] = total_float  # Já formata como número no Excel
            ws[f'E{row_desc}'].number_format = 'R$ #,##0.00'
            ws[f'E{row_desc}'].font = fonte_negrito
            ws[f'E{row_desc}'].alignment = alinhamento_direita
            for row in ws[f'A{row_desc}:E{row_desc}']:
                for cell in row:
                    cell.fill = fill_amarelo
            row_desc += 4

            # Rodapé
            ws.merge_cells(f'D{row_desc}:H{row_desc}')
            ws[f'D{row_desc}'] = f"{self.cidade.get()},               /                  /       "
            ws[f'D{row_desc}'].font = fonte_14
            ws[f'D{row_desc}'].alignment = alinhamento_centro
            row_desc += 2

            ws.merge_cells(f'B{row_desc}:G{row_desc}')
            ws[f'B{row_desc}'] = "__________________________________________________"
            ws[f'B{row_desc}'].font = fonte_14
            ws[f'B{row_desc}'].alignment = alinhamento_centro
            row_desc += 1

            ws.merge_cells(f'B{row_desc}:G{row_desc}')
            ws[f'B{row_desc}'] = self.funcionario_selecionado.get()
            ws[f'B{row_desc}'].font = negrito_14
            ws[f'B{row_desc}'].alignment = alinhamento_centro
            ws[f'B{row_desc}'].fill = fill_amarelo
            row_desc += 1

            # Obter CNPJ/CPF do funcionário selecionado
            cpf_cnpj = ""
            for func in self.funcionarios:
                if func['nome'] == self.funcionario_selecionado.get():
                    cpf_cnpj = func['cnpj']
                    break

            ws.merge_cells(f'D{row_desc}:E{row_desc}')
            ws[f'D{row_desc}'] = f"CPF/CNPJ {cpf_cnpj}"
            ws[f'D{row_desc}'].font = negrito_11
            ws[f'D{row_desc}'].alignment = alinhamento_centro
            ws[f'D{row_desc}'].fill = fill_amarelo

            # Ajustar colunas
            for col in ['A', 'B', 'C', 'D', 'G', 'H']:
                ws.column_dimensions[col].width = 9
            for col in ['E']:
                ws.column_dimensions[col].width = 17
            for col in ['F']:
                ws.column_dimensions[col].width = 13

            # Salvar
            data_atual = datetime.now().strftime("%Y%m%d_%H%M")
            nome_arquivo = f"Recibo_{tipo_recibo}_{self.funcionario_selecionado.get()}_{data_atual}.xlsx"
            wb.save(nome_arquivo)
        
            messagebox.showinfo("Sucesso", f"Recibo gerado com sucesso:\n{nome_arquivo}")
        
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao gerar o recibo:\n{str(e)}")
            print(self.calcular_total)

if __name__ == "__main__":
    root = Tk()
    app = AplicativoRecibo(root)
    root.mainloop()
